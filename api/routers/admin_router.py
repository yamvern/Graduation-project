from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import Response
from datetime import datetime, timezone
import csv
import io
from io import BytesIO

from api.database import get_user_collection, database
from ..models import UserCreate, UserUpdate
from ..security import get_current_admin, get_current_super_admin
from ..security import get_password_hash


def _find_font_path() -> str | None:
    candidates = [
        "C:/Windows/Fonts/arial.ttf",
        "C:/Windows/Fonts/arialuni.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/truetype/noto/NotoSans-Regular.ttf",
    ]
    try:
        from pathlib import Path

        for path in candidates:
            if Path(path).exists():
                return path
    except Exception:
        return None
    return None


def _safe_pdf_text(value: object) -> str:
    text = "" if value is None else str(value)
    try:
        text.encode("latin-1")
        return text
    except Exception:
        return text.encode("latin-1", errors="ignore").decode("latin-1")


def _pdf_write_line(pdf, text: str, line_height: float = 5.0) -> None:
    width = pdf.w - pdf.l_margin - pdf.r_margin
    if width <= 0:
        return
    text = (text or "").replace("\n", " ").strip()
    if text:
        text = (
            text.replace("_", " _ ")
            .replace("/", " / ")
            .replace("\\", " \\ ")
            .replace("-", " - ")
        )
    if not text:
        pdf.ln(line_height)
        return

    words = text.split(" ")
    line = ""
    for word in words:
        candidate = word if not line else f"{line} {word}"
        if pdf.get_string_width(candidate) <= width:
            line = candidate
            continue
        if line:
            pdf.cell(0, line_height, line, ln=1)
            line = ""
        # If the single word is too long, split by chars
        if pdf.get_string_width(word) <= width:
            line = word
        else:
            chunk = ""
            for ch in word:
                if pdf.get_string_width(chunk + ch) <= width:
                    chunk += ch
                else:
                    if chunk:
                        pdf.cell(0, line_height, chunk, ln=1)
                    chunk = ch
            line = chunk
    if line:
        pdf.cell(0, line_height, line, ln=1)


def _pdf_kv(pdf, label: str, value: object, line_height: float = 5.0, label_width: float = 45.0) -> None:
    width = pdf.w - pdf.l_margin - pdf.r_margin
    if width <= 0:
        return
    label_text = (label or "").strip()
    value_text = "" if value is None else str(value)
    value_text = value_text.replace("\n", " ").strip()
    if not label_text:
        _pdf_write_line(pdf, value_text, line_height=line_height)
        return
    # Label column
    pdf.cell(label_width, line_height, label_text, ln=0)
    # Value column
    remaining = max(10.0, width - label_width)
    x = pdf.get_x()
    y = pdf.get_y()
    pdf.multi_cell(remaining, line_height, value_text)
    # Align next line after multi_cell
    pdf.set_xy(pdf.l_margin, max(y + line_height, pdf.get_y()))


def _chart_image_status_pie(data: dict) -> BytesIO | None:
    breakdown = data.get("status_breakdown") or {}
    if not breakdown:
        return None
    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    labels = list(breakdown.keys())
    values = [breakdown[k] for k in labels]
    fig, ax = plt.subplots(figsize=(4.2, 3.0), dpi=150)
    ax.pie(values, labels=labels, autopct="%1.1f%%", startangle=90)
    ax.axis("equal")
    buf = BytesIO()
    fig.tight_layout()
    fig.savefig(buf, format="png")
    plt.close(fig)
    buf.seek(0)
    return buf


def _chart_image_doc_type_bar(data: dict) -> BytesIO | None:
    rows = data.get("by_document_type") or []
    if not rows:
        return None
    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    labels = [r.get("type") or "Unknown" for r in rows]
    values = [r.get("count") or 0 for r in rows]
    fig, ax = plt.subplots(figsize=(6.0, 3.0), dpi=150)
    ax.bar(range(len(values)), values, color="#3b82f6")
    ax.set_xticks(range(len(labels)))
    ax.set_xticklabels(labels, rotation=45, ha="right", fontsize=7)
    ax.set_ylabel("Count")
    fig.tight_layout()
    buf = BytesIO()
    fig.savefig(buf, format="png")
    plt.close(fig)
    buf.seek(0)
    return buf


def _chart_image_failure_bar(data: dict) -> BytesIO | None:
    rows = data.get("failure_reasons") or []
    if not rows:
        return None
    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    labels = [r.get("reason") or "UNKNOWN" for r in rows]
    values = [r.get("count") or 0 for r in rows]
    fig, ax = plt.subplots(figsize=(6.0, 3.0), dpi=150)
    ax.barh(range(len(values)), values, color="#ef4444")
    ax.set_yticks(range(len(labels)))
    ax.set_yticklabels(labels, fontsize=7)
    ax.invert_yaxis()
    ax.set_xlabel("Count")
    fig.tight_layout()
    buf = BytesIO()
    fig.savefig(buf, format="png")
    plt.close(fig)
    buf.seek(0)
    return buf


def _chart_image_time_series(data: dict) -> BytesIO | None:
    rows = data.get("time_series") or []
    if not rows:
        return None
    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    dates = [r.get("date") for r in rows]
    series = {
        "SUCCESS": [r.get("SUCCESS", 0) for r in rows],
        "FAILED": [r.get("FAILED", 0) for r in rows],
        "RUNNING": [r.get("RUNNING", 0) for r in rows],
        "PENDING": [r.get("PENDING", 0) for r in rows],
    }
    fig, ax = plt.subplots(figsize=(7.2, 3.0), dpi=150)
    for name, vals in series.items():
        ax.plot(dates, vals, label=name)
    ax.legend(fontsize=7, ncol=4)
    ax.tick_params(axis="x", labelrotation=45, labelsize=7)
    ax.set_ylabel("Count")
    fig.tight_layout()
    buf = BytesIO()
    fig.savefig(buf, format="png")
    plt.close(fig)
    buf.seek(0)
    return buf


async def _resolve_user_doc(users, user_id: str) -> dict:
    """
    Resolve a user document by id (numeric), username, or email.
    Raises HTTPException 404 if not found.
    """
    uid_val = None
    try:
        uid_val = int(str(user_id).strip())
    except Exception:
        uid_val = None

    user = None
    if uid_val is not None:
        user = await users.find_one({"_id": uid_val})
    if user is None:
        # fallback: try username/email
        if isinstance(user_id, str) and "@" in user_id:
            user = await users.find_one({"email": user_id})
        else:
            user = await users.find_one({"username": user_id})
    if not user:
        raise HTTPException(404, "User not found")
    return user


router = APIRouter(prefix="/api/v1/admin", tags=["Admin"])


def to_public_user(user: dict) -> dict:
    user["_id"] = str(user["_id"])
    user.pop("password", None)
    is_active = user.get("is_active")
    user["is_active"] = True if is_active is None else bool(is_active)
    user["deleted_at"] = user.get("deleted_at")
    return user


def _normalize_header(value: str) -> str:
    return (
        (value or "")
        .strip()
        .lower()
        .replace(" ", "_")
        .replace("-", "_")
        .replace("__", "_")
    )


def _map_headers(headers: list[str]) -> dict[str, int]:
    mapping = {}
    for idx, raw in enumerate(headers):
        key = _normalize_header(raw)
        if key in ("name", "full_name", "full_name_ar", "fullname", "اسم", "الاسم"):
            mapping["name"] = idx
        elif key in ("username", "user_name", "اسم_المستخدم", "اسم_المستخدم_إنجليزي"):
            mapping["username"] = idx
        elif key in ("email", "email_address", "البريد", "البريد_الالكتروني", "البريد_الإلكتروني"):
            mapping["email"] = idx
        elif key in ("password", "pass", "passcode", "كلمة_المرور", "الرمز"):
            mapping["password"] = idx
        elif key in ("role", "user_role", "الدور", "الصلاحية", "النوع"):
            mapping["role"] = idx
    return mapping


# =========================
# Users list (admin + super)
# =========================
@router.get("/users")
async def get_users(admin=Depends(get_current_admin)):
    users = get_user_collection()
    return [to_public_user(u) async for u in users.find({"role": "user"})]


@router.post("/users")
async def create_user(user: UserCreate, admin=Depends(get_current_admin)):
    users = get_user_collection()

    if await users.find_one({"email": user.email}):
        raise HTTPException(status_code=400, detail="Email already registered")
    if user.username and await users.find_one({"username": user.username}):
        raise HTTPException(status_code=400, detail="Username already registered")

    await users.insert_one(
        {
            "name": user.name,
            "username": user.username,
            "email": user.email,
            "password": get_password_hash(user.password),
            "role": "user",
            "is_active": True,
            "deleted_at": None,
        }
    )

    return {"message": "User created"}


@router.get("/users/template")
async def download_users_template(admin=Depends(get_current_admin)):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "users"
    ws.append(["name", "username", "email", "password"])
    ws.append(["Example User", "example_user", "user@example.com", "ChangeMe123"])

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    filename = "user_data_template.xlsx"
    return Response(
        stream.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/users/import")
async def import_users(file: UploadFile = File(...), admin=Depends(get_current_admin)):
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Please upload an .xlsx file")

    from openpyxl import load_workbook

    content = await file.read()
    try:
        wb = load_workbook(BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid Excel file")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Empty Excel file")

    headers = [str(c or "").strip() for c in rows[0]]
    mapping = _map_headers(headers)
    missing = [k for k in ("name", "email", "password") if k not in mapping]
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"Missing required columns: {', '.join(missing)}",
        )

    users = get_user_collection()
    created = 0
    skipped = 0
    errors: list[dict] = []
    seen_emails: set[str] = set()
    seen_usernames: set[str] = set()

    for idx, row in enumerate(rows[1:], start=2):
        if row is None:
            continue

        def _cell(key: str) -> str:
            val = row[mapping[key]] if mapping.get(key) is not None and mapping[key] < len(row) else None
            return "" if val is None else str(val).strip()

        name = _cell("name")
        username = _cell("username") if "username" in mapping else ""
        email = _cell("email").lower()
        password = _cell("password")

        if not name and not email and not password and not username:
            continue
        if not name or not email or not password:
            errors.append({"row": idx, "error": "name, email, password are required"})
            skipped += 1
            continue
        if "@" not in email:
            errors.append({"row": idx, "error": "invalid email"})
            skipped += 1
            continue
        if email in seen_emails:
            errors.append({"row": idx, "error": "duplicate email in file"})
            skipped += 1
            continue
        if username and username in seen_usernames:
            errors.append({"row": idx, "error": "duplicate username in file"})
            skipped += 1
            continue

        if await users.find_one({"email": email}):
            errors.append({"row": idx, "error": "email already exists"})
            skipped += 1
            continue
        if username and await users.find_one({"username": username}):
            errors.append({"row": idx, "error": "username already exists"})
            skipped += 1
            continue

        await users.insert_one(
            {
                "name": name,
                "username": username or None,
                "email": email,
                "password": get_password_hash(password),
                "role": "user",
                "is_active": True,
                "deleted_at": None,
            }
        )
        created += 1
        seen_emails.add(email)
        if username:
            seen_usernames.add(username)

    return {
        "created": created,
        "skipped": skipped,
        "errors": errors[:50],
        "error_count": len(errors),
    }


# =========================
# Edit user data (admin + super)
# =========================
@router.put("/users/{user_id}")
async def update_user(
    user_id: str,
    body: UserUpdate,
    admin=Depends(get_current_admin),
):
    users = get_user_collection()
    user = await _resolve_user_doc(users, user_id)

    # Permission: only the first super admin can edit super_admin users
    if user.get("role") == "super_admin":
        requester_id = str(admin.get("sub"))
        first_sa_id = await _get_first_super_admin_id()
        if requester_id != first_sa_id:
            raise HTTPException(
                403, "Only the primary super admin can edit super admins"
            )

    update_fields = {}
    if body.name is not None:
        update_fields["name"] = body.name
    if body.username is not None:
        # Check uniqueness
        existing = await users.find_one({"username": body.username})
        if existing and existing["_id"] != user["_id"]:
            raise HTTPException(400, "Username already taken")
        update_fields["username"] = body.username
    if body.email is not None:
        existing = await users.find_one({"email": body.email})
        if existing and existing["_id"] != user["_id"]:
            raise HTTPException(400, "Email already registered")
        update_fields["email"] = body.email
    if body.password is not None:
        update_fields["password"] = get_password_hash(body.password)

    if not update_fields:
        raise HTTPException(400, "No fields to update")

    await users.update_one({"_id": user["_id"]}, {"$set": update_fields})
    return {"message": "User updated"}


async def _get_first_super_admin_id() -> str | None:
    """Return the _id of the first (earliest) super_admin in the DB."""
    users = get_user_collection()
    return await users.get_first_super_admin_id()


# =========================
# Admins list (super only)
# =========================
@router.get("/admins")
async def get_admins(super_admin=Depends(get_current_super_admin)):
    users = get_user_collection()
    first_sa_id = await _get_first_super_admin_id()
    result = []
    async for u in users.find({"role": {"$in": ["admin", "super_admin"]}}):
        pub = to_public_user(u)
        pub["is_first_super_admin"] = pub["_id"] == first_sa_id
        result.append(pub)
    return result


@router.post("/admins")
async def create_admin(user: UserCreate, super_admin=Depends(get_current_super_admin)):
    users = get_user_collection()

    if await users.find_one({"email": user.email}):
        raise HTTPException(status_code=400, detail="Email already registered")
    if user.username and await users.find_one({"username": user.username}):
        raise HTTPException(status_code=400, detail="Username already registered")

    await users.insert_one(
        {
            "name": user.name,
            "username": user.username,
            "email": user.email,
            "password": get_password_hash(user.password),
            "role": "admin",
            "is_active": True,
            "deleted_at": None,
        }
    )

    return {"message": "Admin created"}


@router.get("/admins/template")
async def download_admins_template(super_admin=Depends(get_current_super_admin)):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "admins"
    ws.append(["name", "username", "email", "password", "role"])
    ws.append(["Example Admin", "example_admin", "admin@example.com", "ChangeMe123", "admin"])
    ws.append(["Example Super", "super_admin_1", "super@example.com", "ChangeMe123", "super_admin"])

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    filename = "admin_data_template.xlsx"
    return Response(
        stream.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/admins/import")
async def import_admins(
    file: UploadFile = File(...),
    super_admin=Depends(get_current_super_admin),
):
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Please upload an .xlsx file")

    from openpyxl import load_workbook

    content = await file.read()
    try:
        wb = load_workbook(BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid Excel file")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Empty Excel file")

    headers = [str(c or "").strip() for c in rows[0]]
    mapping = _map_headers(headers)
    missing = [k for k in ("name", "email", "password", "role") if k not in mapping]
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"Missing required columns: {', '.join(missing)}",
        )

    users = get_user_collection()
    created = 0
    skipped = 0
    errors: list[dict] = []
    seen_emails: set[str] = set()
    seen_usernames: set[str] = set()

    requester_id = str(super_admin.get("sub"))
    first_sa_id = await _get_first_super_admin_id()
    is_requester_first_sa = requester_id == (first_sa_id or "")

    for idx, row in enumerate(rows[1:], start=2):
        if row is None:
            continue

        def _cell(key: str) -> str:
            val = row[mapping[key]] if mapping.get(key) is not None and mapping[key] < len(row) else None
            return "" if val is None else str(val).strip()

        name = _cell("name")
        username = _cell("username") if "username" in mapping else ""
        email = _cell("email").lower()
        password = _cell("password")
        role = _cell("role").lower()

        if not name and not email and not password and not username and not role:
            continue
        if not name or not email or not password or not role:
            errors.append({"row": idx, "error": "name, email, password, role are required"})
            skipped += 1
            continue
        if "@" not in email:
            errors.append({"row": idx, "error": "invalid email"})
            skipped += 1
            continue
        if role not in ("admin", "super_admin"):
            errors.append({"row": idx, "error": "role must be admin or super_admin"})
            skipped += 1
            continue
        if role == "super_admin" and not is_requester_first_sa:
            errors.append({"row": idx, "error": "only the primary super admin can create super_admin"})
            skipped += 1
            continue
        if email in seen_emails:
            errors.append({"row": idx, "error": "duplicate email in file"})
            skipped += 1
            continue
        if username and username in seen_usernames:
            errors.append({"row": idx, "error": "duplicate username in file"})
            skipped += 1
            continue

        if await users.find_one({"email": email}):
            errors.append({"row": idx, "error": "email already exists"})
            skipped += 1
            continue
        if username and await users.find_one({"username": username}):
            errors.append({"row": idx, "error": "username already exists"})
            skipped += 1
            continue

        await users.insert_one(
            {
                "name": name,
                "username": username or None,
                "email": email,
                "password": get_password_hash(password),
                "role": role,
                "is_active": True,
                "deleted_at": None,
            }
        )
        created += 1
        seen_emails.add(email)
        if username:
            seen_usernames.add(username)

    return {
        "created": created,
        "skipped": skipped,
        "errors": errors[:50],
        "error_count": len(errors),
    }


# =========================
# Promote user → admin
# =========================
@router.put("/users/{user_id}/make-admin")
async def make_admin(user_id: str, super_admin=Depends(get_current_super_admin)):
    users = get_user_collection()

    user = await _resolve_user_doc(users, user_id)

    await users.update_one({"_id": user["_id"]}, {"$set": {"role": "admin"}})

    return {"message": "User promoted to admin"}


# =========================
# Demote admin → user
# =========================
@router.put("/users/{user_id}/remove-admin")
async def remove_admin(user_id: str, super_admin=Depends(get_current_super_admin)):
    users = get_user_collection()
    user = await _resolve_user_doc(users, user_id)

    if user["role"] == "super_admin":
        requester_id = str(super_admin.get("sub"))
        first_sa_id = await _get_first_super_admin_id()
        if requester_id != first_sa_id:
            raise HTTPException(
                403, "Only the primary super admin can demote other super admins"
            )
        if str(user["_id"]) == first_sa_id:
            raise HTTPException(403, "Cannot demote the primary super admin")

    await users.update_one({"_id": user["_id"]}, {"$set": {"role": "user"}})

    return {"message": "Admin removed"}


# =========================
# Suspend / Activate user
# =========================
@router.put("/users/{user_id}/suspend")
async def suspend_user(
    user_id: str,
    admin=Depends(get_current_admin),
):
    users = get_user_collection()
    user = await _resolve_user_doc(users, user_id)
    if user.get("role") == "super_admin":
        raise HTTPException(403, "Cannot suspend super admin")

    await users.update_one({"_id": user["_id"]}, {"$set": {"is_active": False}})
    return {"message": "User suspended"}


@router.put("/users/{user_id}/activate")
async def activate_user(
    user_id: str,
    admin=Depends(get_current_admin),
):
    users = get_user_collection()
    user = await _resolve_user_doc(users, user_id)
    if user.get("role") == "super_admin":
        raise HTTPException(403, "Cannot activate super admin")

    await users.update_one({"_id": user["_id"]}, {"$set": {"is_active": True}})
    return {"message": "User activated"}


# =========================
# Soft delete user
# =========================
@router.delete("/users/{user_id}")
async def soft_delete_user(
    user_id: str,
    admin=Depends(get_current_admin),
):
    users = get_user_collection()
    user = await _resolve_user_doc(users, user_id)
    if user.get("role") == "super_admin":
        raise HTTPException(403, "Cannot delete super admin")

    await users.update_one(
        {"_id": user["_id"]},
        {"$set": {"deleted_at": datetime.now(timezone.utc), "is_active": False}},
    )
    return {"message": "User soft-deleted"}


# =========================
# Analytics summary (admin + super)
# =========================
@router.get("/analytics")
async def get_analytics(
    admin=Depends(get_current_admin),
    date_from: str | None = None,
    date_to: str | None = None,
):
    return await _compute_analytics(date_from=date_from, date_to=date_to)


async def _compute_analytics(date_from: str | None = None, date_to: str | None = None) -> dict:
    if not database.is_connected:
        await database.connect()

    # Build optional date filter for verifications
    date_clause = ""
    values: dict = {}
    if date_from:
        date_clause += " AND v.created_at >= :date_from"
        values["date_from"] = date_from
    if date_to:
        date_clause += " AND v.created_at <= :date_to"
        values["date_to"] = date_to

    async def _count(query: str, vals: dict | None = None) -> int:
        row = await database.fetch_one(query, values=vals)
        return int(row["total"]) if row else 0

    total_users = await _count(
        "SELECT COUNT(*) as total FROM users WHERE role = 'user' AND deleted_at IS NULL"
    )
    total_admins = await _count(
        "SELECT COUNT(*) as total FROM users WHERE role IN ('admin','super_admin') AND deleted_at IS NULL"
    )
    total_verifications = await _count(
        f"SELECT COUNT(*) as total FROM verifications v WHERE 1=1{date_clause}", values
    )
    total_document_types = await _count("SELECT COUNT(*) as total FROM document_types")
    total_audit_logs = await _count("SELECT COUNT(*) as total FROM audit_logs")

    # Status breakdown
    status_rows = await database.fetch_all(
        f"SELECT v.status, COUNT(*) as cnt FROM verifications v WHERE 1=1{date_clause} GROUP BY v.status",
        values=values,
    )
    status_breakdown = {r["status"]: int(r["cnt"]) for r in status_rows}

    # By document type
    type_rows = await database.fetch_all(
        f"""SELECT dt.name as doc_type, COUNT(*) as cnt
            FROM verifications v
            LEFT JOIN document_types dt ON v.document_type_id = dt.id
            WHERE 1=1{date_clause}
            GROUP BY v.document_type_id, dt.name""",
        values=values,
    )
    by_document_type = [
        {"type": r["doc_type"] or "Unknown", "count": int(r["cnt"])} for r in type_rows
    ]

    # Daily time-series (last 30 days if no range)
    ts_clause = (
        date_clause
        if date_clause
        else " AND v.created_at >= DATE_SUB(NOW(), INTERVAL 30 DAY)"
    )
    ts_values = values if date_clause else {}
    time_series = await database.fetch_all(
        f"""SELECT DATE(v.created_at) as day, v.status, COUNT(*) as cnt
            FROM verifications v WHERE 1=1{ts_clause}
            GROUP BY DATE(v.created_at), v.status
            ORDER BY day""",
        values=ts_values,
    )
    daily_data: dict = {}
    for r in time_series:
        day = str(r["day"])
        if day not in daily_data:
            daily_data[day] = {
                "date": day,
                "SUCCESS": 0,
                "FAILED": 0,
                "RUNNING": 0,
                "PENDING": 0,
            }
        daily_data[day][r["status"]] = int(r["cnt"])
    time_series_list = list(daily_data.values())

    # Failure reasons (top 10)
    failure_rows = await database.fetch_all(
        f"""SELECT JSON_EXTRACT(v.result_data, '$.failure_reason_code') as reason, COUNT(*) as cnt
            FROM verifications v
            WHERE v.status = 'FAILED'{date_clause}
            GROUP BY reason ORDER BY cnt DESC LIMIT 10""",
        values=values,
    )
    failure_reasons = [
        {"reason": (r["reason"] or "UNKNOWN").strip('"'), "count": int(r["cnt"])}
        for r in failure_rows
    ]

    # Average processing time (seconds)
    avg_row = await database.fetch_one(
        f"""SELECT AVG(TIMESTAMPDIFF(SECOND, v.start_time, v.end_time)) as avg_sec
            FROM verifications v
            WHERE v.end_time IS NOT NULL AND v.start_time IS NOT NULL{date_clause}""",
        values=values,
    )
    avg_processing_time = round(float(avg_row["avg_sec"] or 0), 1)

    return {
        "total_users": total_users,
        "total_admins": total_admins,
        "total_verifications": total_verifications,
        "total_authentications": total_verifications,
        "total_document_types": total_document_types,
        "total_audit_logs": total_audit_logs,
        "status_breakdown": status_breakdown,
        "by_document_type": by_document_type,
        "time_series": time_series_list,
        "failure_reasons": failure_reasons,
        "avg_processing_time_sec": avg_processing_time,
    }


@router.get("/analytics/export")
async def export_analytics(
    format: str = Query("csv", description="Export format: csv or pdf"),
    date_from: str | None = None,
    date_to: str | None = None,
    admin=Depends(get_current_admin),
):
    fmt = (format or "csv").lower()
    if fmt not in ("csv", "pdf"):
        raise HTTPException(status_code=400, detail="Unsupported export format")

    data = await _compute_analytics(date_from=date_from, date_to=date_to)

    if fmt == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["section", "key", "value"])

        for key in (
            "total_users",
            "total_admins",
            "total_verifications",
            "total_authentications",
            "total_document_types",
            "total_audit_logs",
            "avg_processing_time_sec",
        ):
            writer.writerow(["summary", key, data.get(key)])

        for status, count in (data.get("status_breakdown") or {}).items():
            writer.writerow(["status_breakdown", status, count])

        for row in data.get("by_document_type") or []:
            writer.writerow(["by_document_type", row.get("type"), row.get("count")])

        for row in data.get("failure_reasons") or []:
            writer.writerow(["failure_reasons", row.get("reason"), row.get("count")])

        for row in data.get("time_series") or []:
            writer.writerow(
                [
                    "time_series",
                    row.get("date"),
                    f"SUCCESS={row.get('SUCCESS',0)};FAILED={row.get('FAILED',0)};RUNNING={row.get('RUNNING',0)};PENDING={row.get('PENDING',0)}",
                ]
            )

        filename = f'analytics_export_{datetime.utcnow().strftime("%Y%m%d_%H%M%S")}.csv'
        return Response(
            output.getvalue().encode("utf-8"),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    # PDF
    from fpdf import FPDF

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.add_page()

    font_path = _find_font_path()
    if font_path:
        pdf.add_font("ReportFont", "", font_path, uni=True)
        pdf.set_font("ReportFont", size=10)
        safe = lambda s: "" if s is None else str(s)
    else:
        pdf.set_font("Helvetica", size=10)
        safe = _safe_pdf_text

    pdf.set_font_size(12)
    pdf.cell(0, 8, safe("Analytics Export"), ln=1)
    pdf.set_font_size(10)
    _pdf_kv(pdf, "Export Version", "2026-02-09")
    pdf.ln(1)

    pdf.cell(0, 6, safe("Summary"), ln=1)
    for key in (
        "total_users",
        "total_admins",
        "total_verifications",
        "total_authentications",
        "total_document_types",
        "total_audit_logs",
        "avg_processing_time_sec",
    ):
        _pdf_kv(pdf, key, data.get(key))

    pdf.ln(2)
    pdf.ln(2)
    pdf.cell(0, 6, safe("Status Breakdown"), ln=1)
    for status, count in (data.get("status_breakdown") or {}).items():
        _pdf_kv(pdf, status, count)

    status_pie = _chart_image_status_pie(data)
    if status_pie:
        from PIL import Image

        img = Image.open(status_pie)
        pdf.ln(1)
        pdf.image(img, w=120)

    pdf.ln(2)
    pdf.ln(2)
    pdf.cell(0, 6, safe("By Document Type"), ln=1)
    for row in data.get("by_document_type") or []:
        _pdf_kv(pdf, row.get("type") or "Unknown", row.get("count"))

    doc_bar = _chart_image_doc_type_bar(data)
    if doc_bar:
        from PIL import Image

        img = Image.open(doc_bar)
        pdf.ln(1)
        pdf.image(img, w=190)

    pdf.ln(2)
    pdf.ln(2)
    pdf.cell(0, 6, safe("Failure Reasons"), ln=1)
    for row in data.get("failure_reasons") or []:
        _pdf_kv(pdf, row.get("reason") or "UNKNOWN", row.get("count"))

    failure_bar = _chart_image_failure_bar(data)
    if failure_bar:
        from PIL import Image

        img = Image.open(failure_bar)
        pdf.ln(1)
        pdf.image(img, w=190)

    pdf.ln(2)
    pdf.ln(2)
    pdf.cell(0, 6, safe("Time Series"), ln=1)
    for row in data.get("time_series") or []:
        _pdf_kv(pdf, "Date", row.get("date"))
        pdf.set_font_size(9)
        _pdf_kv(pdf, "SUCCESS", row.get("SUCCESS", 0))
        _pdf_kv(pdf, "FAILED", row.get("FAILED", 0))
        _pdf_kv(pdf, "RUNNING", row.get("RUNNING", 0))
        _pdf_kv(pdf, "PENDING", row.get("PENDING", 0))
        pdf.set_font_size(10)
        pdf.ln(1)

    ts_chart = _chart_image_time_series(data)
    if ts_chart:
        from PIL import Image

        img = Image.open(ts_chart)
        pdf.ln(1)
        pdf.image(img, w=200)

    filename = f'analytics_export_{datetime.utcnow().strftime("%Y%m%d_%H%M%S")}.pdf'
    pdf_bytes = pdf.output(dest="S")
    if isinstance(pdf_bytes, str):
        pdf_bytes = pdf_bytes.encode("latin-1", errors="ignore")
    elif isinstance(pdf_bytes, bytearray):
        pdf_bytes = bytes(pdf_bytes)

    return Response(
        pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
