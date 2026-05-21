from datetime import datetime
from io import BytesIO
import json
from pathlib import Path
from typing import Any, Optional

from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import Response
from fpdf import FPDF
from fpdf.errors import FPDFException
from openpyxl import Workbook

from api.database import get_audit_log_collection
from api.models import AuditLogListResponse
from api.security import get_current_admin

router = APIRouter(
    prefix="/api/admin",
    tags=["Admin - Audit Logs"],
    dependencies=[Depends(get_current_admin)],
)


@router.get("/audit-logs", response_model=AuditLogListResponse)
async def list_audit_logs(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    user_id: Optional[int] = None,
    user_name: Optional[str] = None,
    user_email: Optional[str] = None,
    operation_type: Optional[str] = None,
    status: Optional[str] = None,
    date_from: Optional[datetime] = None,
    date_to: Optional[datetime] = None,
    query: Optional[str] = None,
    sort_by: Optional[str] = Query("created_at"),
    sort_order: Optional[str] = Query("desc"),
):
    collection = get_audit_log_collection()
    filters = {
        "user_id": user_id,
        "user_name": user_name,
        "user_email": user_email,
        "operation_type": operation_type,
        "status": status,
        "date_from": date_from,
        "date_to": date_to,
        "query": query,
    }
    offset = (page - 1) * page_size
    items = await collection.list(
        filters, limit=page_size, offset=offset, sort_by=sort_by, sort_order=sort_order
    )
    total = await collection.count(filters)

    return {
        "total": total,
        "page": page,
        "page_size": page_size,
        "items": items,
    }


def _format_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def _soft_wrap(text: str, chunk: int = 40) -> str:
    """Insert spaces into very long tokens to avoid FPDF line-break errors."""
    if not text:
        return ""
    parts = []
    for token in text.split(" "):
        if len(token) <= chunk:
            parts.append(token)
            continue
        pieces = [token[i : i + chunk] for i in range(0, len(token), chunk)]
        parts.append(" ".join(pieces))
    return " ".join(parts)


def _safe_pdf_text(text: str, chunk: int = 40) -> str:
    if not text:
        return ""
    safe = _soft_wrap(text.replace("\n", " ").replace("\r", " "), chunk=chunk)
    # FPDF core expects latin-1 when not using a unicode font.
    return safe.encode("latin-1", errors="replace").decode("latin-1")


def _find_font_path() -> Optional[str]:
    candidates = [
        Path("C:/Windows/Fonts/arial.ttf"),
        Path("C:/Windows/Fonts/arialuni.ttf"),
        Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
        Path("/usr/share/fonts/truetype/noto/NotoSans-Regular.ttf"),
    ]
    for path in candidates:
        if path.exists():
            return str(path)
    return None


def _build_pdf(items: list[dict[str, Any]]) -> bytes:
    pdf = FPDF(orientation="L", unit="mm", format="A4")
    margin = 12
    pdf.set_auto_page_break(auto=True, margin=margin)
    pdf.set_left_margin(margin)
    pdf.set_right_margin(margin)
    pdf.add_page()

    font_path = _find_font_path()
    if font_path:
        pdf.add_font("AuditFont", "", font_path, uni=True)
        pdf.set_font("AuditFont", size=9)
    else:
        pdf.set_font("Helvetica", size=9)

    def _pdf_kv(label: str, value: Any, line_height: float = 5.0, label_width: float = 35.0) -> None:
        width = pdf.w - pdf.l_margin - pdf.r_margin
        label_text = (label or "").strip()
        value_text = _safe_pdf_text(_format_value(value), chunk=30)
        pdf.cell(label_width, line_height, label_text, ln=0)
        remaining = max(10.0, width - label_width)
        pdf.multi_cell(remaining, line_height, value_text)

    pdf.set_font_size(12)
    pdf.cell(0, 8, "Audit Logs Export", ln=1)
    pdf.set_font_size(9)

    for item in items:
        _pdf_kv("Time", item.get("created_at"))
        _pdf_kv("User", item.get("user_name") or item.get("user_email"))
        _pdf_kv("Role", item.get("user_role"))
        _pdf_kv("Operation", item.get("operation_type"))
        _pdf_kv("Status", item.get("status"))
        _pdf_kv("Module", item.get("module"))
        _pdf_kv("Path", item.get("path"))
        _pdf_kv("File", item.get("file_name"))
        _pdf_kv("Size", item.get("file_size"))
        _pdf_kv("Failure", item.get("failure_reason"))
        pdf.ln(2)

    raw = pdf.output(dest="S")
    if isinstance(raw, (bytes, bytearray)):
        return bytes(raw)
    return str(raw).encode("latin-1", errors="ignore")


def _build_excel(items: list[dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Audit Logs"

    headers = [
        "Time",
        "User",
        "Email",
        "Role",
        "Operation",
        "Status",
        "Module",
        "Path",
        "Method",
        "File Name",
        "File Ext",
        "File Size",
        "Failure Reason",
        "Extra Data",
    ]
    ws.append(headers)

    for item in items:
        ws.append(
            [
                _format_value(item.get("created_at")),
                _format_value(item.get("user_name")),
                _format_value(item.get("user_email")),
                _format_value(item.get("user_role")),
                _format_value(item.get("operation_type")),
                _format_value(item.get("status")),
                _format_value(item.get("module")),
                _format_value(item.get("path")),
                _format_value(item.get("method")),
                _format_value(item.get("file_name")),
                _format_value(item.get("file_ext")),
                _format_value(item.get("file_size")),
                _format_value(item.get("failure_reason")),
                _format_value(item.get("extra_data")),
            ]
        )

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream.read()


@router.get("/audit-logs/export")
async def export_audit_logs(format: str = "pdf"):
    collection = get_audit_log_collection()
    items = await collection.list_all({})

    if format.lower() == "xlsx":
        content = _build_excel(items)
        filename = f"audit_logs_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return Response(
            content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    if format.lower() == "pdf":
        content = _build_pdf(items)
        filename = f"audit_logs_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.pdf"
        return Response(
            content,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    raise HTTPException(status_code=400, detail="Unsupported export format")
